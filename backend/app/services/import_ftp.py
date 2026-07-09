"""
Import FTP des fichiers réceptions et codes-barres.

Réceptions : <N°EN>_<CodeMagasin>_<CodeFournisseur>_<NumFacture>.xlsx
  - Un fichier = une réception indépendante
  - Déduplication par nom de fichier (source_filename)
  - Colonnes Excel : A=NUMREV, B=Société, C=Etablissement, D=Code fournisseur,
                     E=Fournisseur, F=N° Facture, G=Réf interne, H=Réf fournisseur,
                     I=Désignation, J=Qté attendue

Codes-barres : fichier unique (remplacement intelligent).
"""
import os
import re
import io
from datetime import datetime, timezone

import openpyxl
from loguru import logger
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.database import AsyncSessionLocal
from app.models.models import (
    Article, CodeBarre, ImportLog, LigneReception,
    Magasin, Reception, SourceCodeBarre, StatutImport,
    StatutReception, TypeImport,
)

settings = get_settings()

# Format : EN32600113_PDGMO_F01COU00_FC2657330B.xlsx
RECEPTION_FILENAME_RE = re.compile(
    r"^(?P<numero_en>[^_]+)_(?P<magasin_code>[^_]+)_(?P<code_fournisseur>[^_]+)_(?P<num_facture>.+)\.xlsx$",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Connecteur FTP/SFTP
# ---------------------------------------------------------------------------

def _make_ftp_client():
    if settings.FTP_USE_SFTP:
        import paramiko
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(
            hostname=settings.FTP_HOST,
            port=settings.FTP_PORT,
            username=settings.FTP_USER,
            password=settings.FTP_PASSWORD,
            timeout=30,
        )
        return ssh.open_sftp(), ssh
    else:
        import ftplib

        class _FTPTLSResuming(ftplib.FTP_TLS):
            def ntransfercmd(self, cmd, rest=None):
                conn, size = ftplib.FTP.ntransfercmd(self, cmd, rest)
                if self._prot_p:
                    conn = self.context.wrap_socket(
                        conn,
                        server_hostname=self.host,
                        session=self.sock.session,
                    )
                return conn, size

        ftp = _FTPTLSResuming()
        ftp.connect(settings.FTP_HOST, settings.FTP_PORT, timeout=30)
        ftp.login(settings.FTP_USER, settings.FTP_PASSWORD)
        ftp.prot_p()
        return ftp, None


def _close(client, ssh):
    try:
        client.quit() if hasattr(client, 'quit') else client.close()
    except Exception:
        pass
    if ssh:
        try:
            ssh.close()
        except Exception:
            pass


def _list_files(path: str) -> list[str]:
    client, ssh = _make_ftp_client()
    try:
        if hasattr(client, "listdir"):  # SFTP
            return client.listdir(path)
        else:  # FTP
            return [os.path.basename(f) for f in client.nlst(path)]
    finally:
        _close(client, ssh)


def _read_file(remote_path: str) -> bytes:
    client, ssh = _make_ftp_client()
    try:
        if hasattr(client, "open"):  # SFTP
            with client.open(remote_path, "rb") as f:
                return f.read()
        else:  # FTP
            buf = io.BytesIO()
            client.retrbinary(f"RETR {remote_path}", buf.write)
            return buf.getvalue()
    finally:
        _close(client, ssh)


def _ftp_makedirs(client, path: str):
    """Crée récursivement les répertoires FTP manquants."""
    parts = [p for p in path.split("/") if p]
    current = ""
    for part in parts:
        current += "/" + part
        try:
            client.mkd(current)
        except Exception:
            pass  # existe déjà ou erreur ignorée


def _move_file(src: str, dst_dir: str, filename: str):
    client, ssh = _make_ftp_client()
    try:
        if hasattr(client, "listdir"):  # SFTP
            try:
                client.mkdir(dst_dir)
            except OSError:
                pass
            client.rename(src, f"{dst_dir}/{filename}")
        else:  # FTP
            _ftp_makedirs(client, dst_dir)
            client.rename(src, f"{dst_dir}/{filename}")
    finally:
        _close(client, ssh)


# ---------------------------------------------------------------------------
# Import réceptions
# ---------------------------------------------------------------------------

async def import_receptions(db: AsyncSession) -> ImportLog:
    log = ImportLog(type=TypeImport.receptions, statut=StatutImport.en_cours)
    db.add(log)
    await db.flush()

    if not settings.FTP_HOST:
        log.statut = StatutImport.erreur
        log.message_erreur = "FTP_HOST non configuré"
        log.ended_at = datetime.now(timezone.utc)
        return log

    try:
        ftp_path = settings.FTP_PATH_RECEPTIONS
        files = [f for f in _list_files(ftp_path) if f.lower().endswith(".xlsx")]
        logger.info(f"Import réceptions : {len(files)} fichier(s) trouvé(s)")

        lignes_traitees = 0
        lignes_erreur = 0

        for filename in files:
            m = RECEPTION_FILENAME_RE.match(filename)
            if not m:
                logger.warning(f"Nom de fichier non reconnu (attendu EN_MAGASIN_FOURN_FACTURE.xlsx) : {filename}")
                lignes_erreur += 1
                continue

            numero_en = m.group("numero_en")
            magasin_code = m.group("magasin_code")
            code_fournisseur = m.group("code_fournisseur")
            num_facture = m.group("num_facture")

            # Déduplication : fichier déjà traité si au moins une réception non-ancienne existe
            r_existing = await db.execute(
                select(Reception).where(
                    Reception.source_filename.like(f"{filename}%"),
                    Reception.statut != StatutReception.ancien,
                )
            )
            if r_existing.scalar_one_or_none():
                logger.info(f"Fichier {filename} déjà importé, ignoré")
                _archive_file(ftp_path, filename)
                continue

            r_mag = await db.execute(select(Magasin).where(Magasin.code == magasin_code))
            magasin = r_mag.scalar_one_or_none()
            if not magasin:
                logger.warning(f"Magasin '{magasin_code}' introuvable pour {filename}")
                lignes_erreur += 1
                continue

            try:
                content = _read_file(f"{ftp_path}/{filename}")
                nb = await _process_reception_file(
                    db, content, numero_en, magasin, code_fournisseur, num_facture, filename
                )
                lignes_traitees += nb
                logger.info(f"{filename} importé : {nb} ligne(s)")
            except Exception as e:
                logger.error(f"Erreur traitement {filename}: {e}")
                lignes_erreur += 1
                continue

            _archive_file(ftp_path, filename)

        log.statut = StatutImport.succes
        log.lignes_traitees = lignes_traitees
        log.lignes_erreur = lignes_erreur

    except Exception as e:
        log.statut = StatutImport.erreur
        log.message_erreur = str(e)
        logger.error(f"Import réceptions échoué : {e}")

    log.ended_at = datetime.now(timezone.utc)
    return log


def _archive_file(ftp_path: str, filename: str):
    now = datetime.now(timezone.utc)
    archive_dir = f"{ftp_path}/archive/{now.strftime('%Y-%m')}"
    try:
        _move_file(f"{ftp_path}/{filename}", archive_dir, filename)
    except Exception as e:
        logger.warning(f"Impossible d'archiver {filename}: {e}")


async def _process_reception_file(
    db: AsyncSession,
    content: bytes,
    numero_en: str,
    magasin: Magasin,
    code_fournisseur: str,
    num_facture: str,
    filename: str,
) -> int:
    """
    Traite un fichier Excel de réception.

    Structure des colonnes (ligne d'en-tête en ligne 1, données à partir de la ligne 2) :
      A (0) : NUMREV             - N° EN (par ligne — peut différer du nom de fichier)
      B (1) : Société
      C (2) : Etablissement      - code magasin
      D (3) : Code fournisseur
      E (4) : Fournisseur        - nom du fournisseur
      F (5) : N° Facture Fournisseur
      G (6) : Réf interne        - référence article
      H (7) : Réf fournisseur
      I (8) : Désignation
      J (9) : Qté attendue

    Un fichier peut contenir plusieurs EN (une commande fournisseur = plusieurs ENs).
    On crée une réception distincte par EN trouvé dans la colonne A.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return 0

    # Lire le nom du fournisseur depuis la première ligne de données (col E, index 4)
    fournisseur_nom = code_fournisseur
    for row in rows:
        if row and len(row) > 4 and row[4]:
            fournisseur_nom = str(row[4]).strip()
            break

    # Grouper les lignes par numéro EN (colonne A).
    # Si la colonne A est vide ou absente, on utilise le numéro EN du nom de fichier.
    from collections import defaultdict
    rows_by_en: dict[str, list] = defaultdict(list)
    for row in rows:
        if not row or not any(row):
            continue
        en_in_row = str(row[0]).strip() if row[0] else None
        # L'ERP écrit "EN32600148.00001" en col A — on ne garde que la partie avant le point
        if en_in_row:
            en_key = en_in_row.split(".")[0]
        else:
            en_key = numero_en
        rows_by_en[en_key].append(row)

    nb_total = 0

    for en_key, en_rows in rows_by_en.items():
        # Déduplication : si une réception non-ancienne avec ce filename+EN existe, on ignore
        source_key = f"{filename}#{en_key}"
        r_existing = await db.execute(
            select(Reception).where(
                Reception.source_filename == source_key,
                Reception.statut != StatutReception.ancien,
            )
        )
        if r_existing.scalar_one_or_none():
            logger.info(f"Réception {en_key} du fichier {filename} déjà importée, ignorée")
            continue

        reception = Reception(
            numero_en=en_key,
            magasin_id=magasin.id,
            code_fournisseur=code_fournisseur,
            fournisseur_nom=fournisseur_nom,
            num_facture_fournisseur=num_facture,
            source_filename=source_key,
            statut=StatutReception.en_cours,
            saisie_aveugle=True,
        )
        db.add(reception)
        await db.flush()

        for row in en_rows:
            try:
                reference_interne = str(row[6]).strip() if len(row) > 6 and row[6] else None
                reference_fournisseur = str(row[7]).strip() if len(row) > 7 and row[7] else None
                designation = str(row[8]).strip() if len(row) > 8 and row[8] else "Sans désignation"
                # Qté : Excel peut retourner un float (ex: 60.0), on convertit en int
                qte_raw = row[9] if len(row) > 9 else None
                quantite_attendue = int(float(str(qte_raw))) if qte_raw is not None else None
            except (ValueError, IndexError):
                continue

            if not reference_interne:
                continue

            r = await db.execute(select(Article).where(Article.reference_interne == reference_interne))
            article = r.scalar_one_or_none()
            if not article:
                article = Article(reference_interne=reference_interne, designation=designation)
                db.add(article)
                await db.flush()

            ligne = LigneReception(
                reception_id=reception.id,
                article_id=article.id,
                reference_interne=reference_interne,
                reference_fournisseur=reference_fournisseur,
                designation=designation,
                quantite_attendue=quantite_attendue,
            )
            db.add(ligne)
            nb_total += 1

        await db.flush()
        logger.info(f"Réception {en_key} créée depuis {filename} : {len(en_rows)} ligne(s)")

    return nb_total


# ---------------------------------------------------------------------------
# Import codes-barres (remplacement intelligent)
# ---------------------------------------------------------------------------

async def import_codes_barres(db: AsyncSession) -> ImportLog:
    log = ImportLog(type=TypeImport.codes_barres, statut=StatutImport.en_cours)
    db.add(log)
    await db.flush()

    if not settings.FTP_HOST:
        log.statut = StatutImport.erreur
        log.message_erreur = "FTP_HOST non configuré"
        log.ended_at = datetime.now(timezone.utc)
        return log

    try:
        ftp_path = settings.FTP_PATH_CODES_BARRES
        files = [f for f in _list_files(ftp_path) if f.lower().endswith(".xlsx")]

        if not files:
            log.statut = StatutImport.succes
            log.message_erreur = "Aucun fichier codes-barres trouvé"
            log.ended_at = datetime.now(timezone.utc)
            return log

        filename = files[-1]
        log.fichier_nom = filename
        content = _read_file(f"{ftp_path}/{filename}")

        nb_traites, nb_erreurs = await _process_codes_barres_file(db, content)
        log.statut = StatutImport.succes
        log.lignes_traitees = nb_traites
        log.lignes_erreur = nb_erreurs

    except Exception as e:
        log.statut = StatutImport.erreur
        log.message_erreur = str(e)
        logger.error(f"Import codes-barres échoué : {e}")

    log.ended_at = datetime.now(timezone.utc)
    return log


async def _process_codes_barres_file(db: AsyncSession, content: bytes) -> tuple[int, int]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    file_data: dict[str, list[str]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        ref = str(row[0]).strip()
        code = str(row[1]).strip()
        if ref and code:
            file_data.setdefault(ref, []).append(code)

    all_file_codes = {code for codes in file_data.values() for code in codes}

    nb_traites = 0
    nb_erreurs = 0
    seen_codes: set[str] = set()
    for ref, codes in file_data.items():
        r = await db.execute(select(Article).where(Article.reference_interne == ref))
        article = r.scalar_one_or_none()
        if not article:
            article = Article(reference_interne=ref, designation=ref)
            db.add(article)
            await db.flush()
            logger.info(f"Article {ref} créé automatiquement via import codes-barres")

        for code in codes:
            if code in seen_codes:
                logger.warning(f"Code-barre dupliqué dans le fichier, ignoré : {code}")
                continue
            seen_codes.add(code)
            r2 = await db.execute(select(CodeBarre).where(CodeBarre.code == code))
            existing = r2.scalar_one_or_none()
            if existing:
                if existing.source == SourceCodeBarre.import_:
                    existing.article_id = article.id
            else:
                cb = CodeBarre(
                    article_id=article.id,
                    code=code,
                    source=SourceCodeBarre.import_,
                )
                db.add(cb)
            nb_traites += 1

    await db.flush()

    r = await db.execute(
        select(CodeBarre).where(CodeBarre.source == SourceCodeBarre.import_)
    )
    for cb in r.scalars().all():
        if cb.code not in all_file_codes:
            await db.delete(cb)

    await db.flush()
    return nb_traites, nb_erreurs


# ---------------------------------------------------------------------------
# Point d'entrée unique appelé par le scheduler
# ---------------------------------------------------------------------------

async def run_all_imports():
    async with AsyncSessionLocal() as db:
        try:
            logger.info("=== Démarrage import FTP ===")
            log_rec = await import_receptions(db)
            log_cb = await import_codes_barres(db)
            await db.commit()
            logger.info(
                f"Import terminé — Réceptions: {log_rec.statut.value}, "
                f"Codes-barres: {log_cb.statut.value}"
            )
        except Exception as e:
            await db.rollback()
            logger.error(f"Erreur import global : {e}")
